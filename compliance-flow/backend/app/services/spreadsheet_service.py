"""Spreadsheet service for parsing, exporting, and transforming data."""

import csv
import io
import os
from typing import Any, Dict, List, Optional


class SpreadsheetService:
    """Service for spreadsheet operations: parse, export, and transform."""

    async def parse_file(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> dict:
        """Parse a spreadsheet file and return columns with preview rows.

        Args:
            file_path: Path to the spreadsheet file (csv or xlsx).
            sheet_name: Optional sheet name for xlsx files.

        Returns:
            Dictionary with columns, preview_rows, and total_rows.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            return await self._parse_csv(file_path)
        elif ext in (".xlsx", ".xls"):
            return await self._parse_xlsx(file_path, sheet_name)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    async def _parse_csv(self, file_path: str) -> dict:
        """Parse a CSV file."""
        rows: List[Dict[str, Any]] = []

        with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames or []

            for row in reader:
                rows.append(dict(row))

        preview_rows = rows[:100]

        return {
            "columns": list(columns),
            "preview_rows": preview_rows,
            "total_rows": len(rows),
        }

    async def _parse_xlsx(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> dict:
        """Parse an XLSX file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for XLSX support. "
                "Install it with: pip install openpyxl"
            )

        wb = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # First row is the header
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return {"columns": [], "preview_rows": [], "total_rows": 0}

        columns = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(header_row)]

        rows: List[Dict[str, Any]] = []
        for row_values in rows_iter:
            row_dict = {}
            for i, value in enumerate(row_values):
                col_name = columns[i] if i < len(columns) else f"Column_{i}"
                row_dict[col_name] = value
            rows.append(row_dict)

        wb.close()

        preview_rows = rows[:100]

        return {
            "columns": columns,
            "preview_rows": preview_rows,
            "total_rows": len(rows),
        }

    async def export_data(
        self, data: List[Dict[str, Any]], format: str, output_path: str
    ) -> str:
        """Export data to a spreadsheet file.

        Args:
            data: List of row dictionaries to export.
            format: Output format ('csv' or 'xlsx').
            output_path: Path to write the output file.

        Returns:
            The output file path on success.
        """
        if format == "csv":
            return await self._export_csv(data, output_path)
        elif format == "xlsx":
            return await self._export_xlsx(data, output_path)
        else:
            raise ValueError(f"Unsupported export format: {format}. Use 'csv' or 'xlsx'.")

    async def _export_csv(
        self, data: List[Dict[str, Any]], output_path: str
    ) -> str:
        """Export data to CSV."""
        if not data:
            # Write empty file
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                pass
            return output_path

        columns = list(data[0].keys())

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)

        return output_path

    async def _export_xlsx(
        self, data: List[Dict[str, Any]], output_path: str
    ) -> str:
        """Export data to XLSX using openpyxl."""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for XLSX export. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active

        if not data:
            wb.save(output_path)
            wb.close()
            return output_path

        columns = list(data[0].keys())

        # Write header
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

        wb.save(output_path)
        wb.close()

        return output_path

    async def transform(
        self, data: List[Dict[str, Any]], operations: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Apply a sequence of transformations to data.

        Args:
            data: List of row dictionaries.
            operations: List of operations, each with 'type', 'column', and optional params.

        Supported operation types:
            - filter: Keep rows where column equals value.
            - sort: Sort rows by column (direction: 'asc' or 'desc').
            - aggregate: Group by column and count rows per group.

        Returns:
            Transformed list of row dictionaries.
        """
        result = list(data)

        for op in operations:
            op_type = op.get("type")
            column = op.get("column")

            if not column:
                raise ValueError("Each operation must specify a 'column'.")

            if op_type == "filter":
                value = op.get("value")
                result = [row for row in result if row.get(column) == value]

            elif op_type == "sort":
                direction = op.get("direction", "asc")
                reverse = direction == "desc"
                result = sorted(
                    result,
                    key=lambda row: (row.get(column) is None, row.get(column)),
                    reverse=reverse,
                )

            elif op_type == "aggregate":
                groups: Dict[Any, int] = {}
                for row in result:
                    key = row.get(column)
                    groups[key] = groups.get(key, 0) + 1
                result = [
                    {column: key, "count": count}
                    for key, count in groups.items()
                ]

            else:
                raise ValueError(f"Unsupported operation type: {op_type}")

        return result
