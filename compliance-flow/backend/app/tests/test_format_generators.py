"""
Tests for report_generator.py — Checkpoint 2: Format Generators.

TDD: Tests written BEFORE implementation of format conversion functions.
Tests cover: generate_markdown, generate_pdf, generate_docx, generate_xlsx.
"""

import io
import pytest
from typing import Dict, Any


# ---------------------------------------------------------------------------
# Tests for generate_markdown (pure function)
# ---------------------------------------------------------------------------


class TestGenerateMarkdown:
    """Tests for Markdown output generation."""

    def test_includes_title(self) -> None:
        """Output starts with the report title as an H1 header."""
        from app.services.report_generator import generate_markdown

        result: bytes = generate_markdown(content="Body text here.", title="My Report")
        text: str = result.decode("utf-8")
        assert text.startswith("# My Report")

    def test_includes_content(self) -> None:
        """Output contains the provided content after the title."""
        from app.services.report_generator import generate_markdown

        result: bytes = generate_markdown(content="## Findings\n\nNo gaps.", title="Report")
        text: str = result.decode("utf-8")
        assert "## Findings" in text
        assert "No gaps." in text

    def test_empty_content(self) -> None:
        """Empty content produces valid output without crash."""
        from app.services.report_generator import generate_markdown

        result: bytes = generate_markdown(content="", title="Empty")
        assert isinstance(result, bytes)
        assert len(result) > 0


# ---------------------------------------------------------------------------
# Tests for generate_pdf (pure function)
# ---------------------------------------------------------------------------


class TestGeneratePdf:
    """Tests for PDF output generation."""

    def test_valid_pdf_header(self) -> None:
        """Output starts with the PDF magic bytes."""
        from app.services.report_generator import generate_pdf

        result: bytes = generate_pdf(content="Test content", title="Test")
        assert result[:5] == b"%PDF-"

    def test_empty_content(self) -> None:
        """Empty content still produces a valid PDF."""
        from app.services.report_generator import generate_pdf

        result: bytes = generate_pdf(content="", title="Empty")
        assert result[:5] == b"%PDF-"
        assert len(result) > 100  # Non-trivial PDF size

    def test_multiline_content(self) -> None:
        """Multi-line markdown content produces valid PDF without crash."""
        from app.services.report_generator import generate_pdf

        content: str = "# Section 1\n\nParagraph one.\n\n## Section 2\n\n- Item A\n- Item B"
        result: bytes = generate_pdf(content=content, title="Multi")
        assert result[:5] == b"%PDF-"


# ---------------------------------------------------------------------------
# Tests for generate_docx (pure function)
# ---------------------------------------------------------------------------


class TestGenerateDocx:
    """Tests for DOCX output generation."""

    def test_valid_zip_format(self) -> None:
        """DOCX is a ZIP file — output starts with PK magic bytes."""
        from app.services.report_generator import generate_docx

        result: bytes = generate_docx(content="Test", title="Test")
        assert result[:2] == b"PK"

    def test_contains_title(self) -> None:
        """Extracted text from the DOCX contains the title."""
        from app.services.report_generator import generate_docx
        from docx import Document

        result: bytes = generate_docx(content="Body text.", title="My Title")
        doc = Document(io.BytesIO(result))
        full_text: str = "\n".join(p.text for p in doc.paragraphs)
        assert "My Title" in full_text

    def test_empty_content(self) -> None:
        """Empty content still produces a valid DOCX."""
        from app.services.report_generator import generate_docx

        result: bytes = generate_docx(content="", title="Empty")
        assert result[:2] == b"PK"


# ---------------------------------------------------------------------------
# Tests for generate_xlsx (pure function)
# ---------------------------------------------------------------------------


class TestGenerateXlsx:
    """Tests for XLSX output generation."""

    def test_two_sheets(self) -> None:
        """Workbook contains exactly two sheets: Summary and Data."""
        from app.services.report_generator import generate_xlsx
        from openpyxl import load_workbook

        result: bytes = generate_xlsx(
            content="Summary text", title="Report", input_data={"key": "value"}
        )
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2
        assert "Summary" in wb.sheetnames
        assert "Data" in wb.sheetnames

    def test_data_sheet_content(self) -> None:
        """Data sheet contains input data values."""
        from app.services.report_generator import generate_xlsx
        from openpyxl import load_workbook

        data: Dict[str, Any] = {"status": "pass", "score": 95}
        result: bytes = generate_xlsx(content="Summary", title="Report", input_data=data)
        wb = load_workbook(io.BytesIO(result))
        data_sheet = wb["Data"]
        # Collect all cell values
        values = [cell.value for row in data_sheet.iter_rows() for cell in row if cell.value]
        assert "status" in values or "pass" in values

    def test_empty_data(self) -> None:
        """Empty input_data produces valid XLSX without crash."""
        from app.services.report_generator import generate_xlsx

        result: bytes = generate_xlsx(content="Summary", title="Report", input_data={})
        assert result[:2] == b"PK"  # XLSX is a ZIP file
