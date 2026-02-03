"""
Output API Routes
Handles spreadsheet exports, email sending, and Telegram messaging
"""

from fastapi import APIRouter, HTTPException, BackgroundTasks
from pydantic import BaseModel, Field
from typing import Optional, Dict, Any, List

# Try to import EmailStr, fallback to str if email-validator not installed
try:
    from pydantic import EmailStr
except ImportError:
    EmailStr = str  # type: ignore
import csv
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import asyncio
import aiohttp
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/outputs", tags=["outputs"])

# =====================
# Spreadsheet Models
# =====================

class SpreadsheetExportRequest(BaseModel):
    """Export data to spreadsheet format"""
    data: List[Dict[str, Any]] = Field(..., description="Data rows to export")
    format: str = Field("csv", description="Export format: csv, xlsx")
    filename: Optional[str] = Field(None, description="Output filename")
    include_headers: bool = Field(True, description="Include column headers")

class SpreadsheetExportResponse(BaseModel):
    success: bool
    file_content: str = Field(..., description="Base64 encoded file content")
    filename: str
    mime_type: str
    row_count: int

# =====================
# Email Models
# =====================

class EmailConfig(BaseModel):
    """SMTP server configuration"""
    smtp_host: str = Field(..., description="SMTP server hostname")
    smtp_port: int = Field(587, description="SMTP server port (default: 587 for TLS)")
    smtp_username: str = Field(..., description="SMTP username/email")
    smtp_password: str = Field(..., description="SMTP password")
    use_tls: bool = Field(True, description="Use TLS encryption")

class EmailSendRequest(BaseModel):
    """Send email message"""
    config: EmailConfig
    to_email: EmailStr = Field(..., description="Recipient email address")
    subject: str = Field(..., description="Email subject")
    body: str = Field(..., description="Email body (plain text or HTML)")
    body_type: str = Field("html", description="Body type: plain or html")
    from_name: Optional[str] = Field(None, description="Sender display name")
    attachments: Optional[List[Dict[str, str]]] = Field(None, description="Email attachments")

class EmailSendResponse(BaseModel):
    success: bool
    message: str
    message_id: Optional[str] = None

# =====================
# Telegram Models
# =====================

class TelegramConfig(BaseModel):
    """Telegram bot configuration"""
    bot_token: str = Field(..., description="Telegram bot token from @BotFather")

class TelegramSendRequest(BaseModel):
    """Send Telegram message"""
    config: TelegramConfig
    chat_id: str = Field(..., description="Chat ID or @channel_username")
    text: str = Field(..., description="Message text (supports Markdown)")
    parse_mode: str = Field("Markdown", description="Parse mode: Markdown or HTML")
    disable_notification: bool = Field(False, description="Send silently")

class TelegramSendResponse(BaseModel):
    success: bool
    message: str
    message_id: Optional[int] = None

# =====================
# Spreadsheet Endpoints
# =====================

@router.post("/spreadsheet/export/csv", response_model=SpreadsheetExportResponse)
async def export_to_csv(request: SpreadsheetExportRequest):
    """
    Export data to CSV format

    Converts a list of dictionaries to CSV format with optional headers.
    Returns base64-encoded CSV content for download.
    """
    try:
        if not request.data:
            raise HTTPException(status_code=400, detail="No data provided for export")

        # Create CSV in memory
        output = io.StringIO()

        # Get all unique keys from all rows (for header)
        all_keys = set()
        for row in request.data:
            all_keys.update(row.keys())
        fieldnames = sorted(all_keys)

        writer = csv.DictWriter(output, fieldnames=fieldnames)

        if request.include_headers:
            writer.writeheader()

        for row in request.data:
            writer.writerow(row)

        # Get CSV content
        csv_content = output.getvalue()

        # Encode to base64 for transmission
        import base64
        file_content_b64 = base64.b64encode(csv_content.encode()).decode()

        filename = request.filename or "export.csv"
        if not filename.endswith('.csv'):
            filename += '.csv'

        return SpreadsheetExportResponse(
            success=True,
            file_content=file_content_b64,
            filename=filename,
            mime_type="text/csv",
            row_count=len(request.data)
        )

    except Exception as e:
        logger.error(f"CSV export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.post("/spreadsheet/export/xlsx", response_model=SpreadsheetExportResponse)
async def export_to_xlsx(request: SpreadsheetExportRequest):
    """
    Export data to Excel (XLSX) format

    Requires openpyxl package to be installed.
    Returns base64-encoded XLSX content for download.
    """
    try:
        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail="Excel export not available. Install openpyxl package."
            )

        if not request.data:
            raise HTTPException(status_code=400, detail="No data provided for export")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"

        # Get all unique keys
        all_keys = set()
        for row in request.data:
            all_keys.update(row.keys())
        fieldnames = sorted(all_keys)

        # Write headers
        if request.include_headers:
            for col_idx, key in enumerate(fieldnames, start=1):
                ws.cell(row=1, column=col_idx, value=key)

        # Write data
        start_row = 2 if request.include_headers else 1
        for row_idx, data_row in enumerate(request.data, start=start_row):
            for col_idx, key in enumerate(fieldnames, start=1):
                value = data_row.get(key, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Encode to base64
        import base64
        file_content_b64 = base64.b64encode(output.read()).decode()

        filename = request.filename or "export.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        return SpreadsheetExportResponse(
            success=True,
            file_content=file_content_b64,
            filename=filename,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            row_count=len(request.data)
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

# =====================
# Email Endpoints
# =====================

@router.post("/email/send", response_model=EmailSendResponse)
async def send_email(request: EmailSendRequest, background_tasks: BackgroundTasks):
    """
    Send email via SMTP

    Supports both plain text and HTML emails with optional attachments.
    Uses SMTP with TLS for secure transmission.
    """
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"{request.from_name} <{request.config.smtp_username}>" if request.from_name else request.config.smtp_username
        msg['To'] = request.to_email
        msg['Subject'] = request.subject

        # Add body
        if request.body_type == 'html':
            msg.attach(MIMEText(request.body, 'html'))
        else:
            msg.attach(MIMEText(request.body, 'plain'))

        # Add attachments if provided
        if request.attachments:
            for attachment in request.attachments:
                # Decode base64 attachment
                import base64
                file_data = base64.b64decode(attachment['content'])
                part = MIMEApplication(file_data, Name=attachment['filename'])
                part['Content-Disposition'] = f'attachment; filename="{attachment["filename"]}"'
                msg.attach(part)

        # Connect and send
        try:
            if request.config.use_tls:
                server = smtplib.SMTP(request.config.smtp_host, request.config.smtp_port)
                server.starttls()
            else:
                server = smtplib.SMTP(request.config.smtp_host, request.config.smtp_port)

            server.login(request.config.smtp_username, request.config.smtp_password)
            server.send_message(msg)
            server.quit()

            return EmailSendResponse(
                success=True,
                message=f"Email sent successfully to {request.to_email}",
                message_id=msg['Message-ID']
            )

        except smtplib.SMTPException as smtp_err:
            logger.error(f"SMTP error: {smtp_err}")
            return EmailSendResponse(
                success=False,
                message=f"SMTP error: {str(smtp_err)}"
            )

    except Exception as e:
        logger.error(f"Email send failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")

@router.post("/email/test", response_model=EmailSendResponse)
async def test_email_config(config: EmailConfig):
    """
    Test SMTP configuration

    Verifies SMTP credentials and connection without sending a message.
    """
    try:
        if config.use_tls:
            server = smtplib.SMTP(config.smtp_host, config.smtp_port, timeout=10)
            server.starttls()
        else:
            server = smtplib.SMTP(config.smtp_host, config.smtp_port, timeout=10)

        server.login(config.smtp_username, config.smtp_password)
        server.quit()

        return EmailSendResponse(
            success=True,
            message="SMTP configuration is valid"
        )

    except smtplib.SMTPAuthenticationError:
        return EmailSendResponse(
            success=False,
            message="Authentication failed. Check username and password."
        )
    except smtplib.SMTPException as smtp_err:
        return EmailSendResponse(
            success=False,
            message=f"SMTP error: {str(smtp_err)}"
        )
    except Exception as e:
        logger.error(f"Email config test failed: {e}")
        return EmailSendResponse(
            success=False,
            message=f"Connection failed: {str(e)}"
        )

# =====================
# Telegram Endpoints
# =====================

@router.post("/telegram/send", response_model=TelegramSendResponse)
async def send_telegram_message(request: TelegramSendRequest):
    """
    Send message via Telegram Bot API

    Sends text messages to a Telegram chat or channel.
    Supports Markdown and HTML formatting.
    """
    try:
        url = f"https://api.telegram.org/bot{request.config.bot_token}/sendMessage"

        payload = {
            "chat_id": request.chat_id,
            "text": request.text,
            "parse_mode": request.parse_mode,
            "disable_notification": request.disable_notification
        }

        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=payload) as response:
                result = await response.json()

                if result.get('ok'):
                    return TelegramSendResponse(
                        success=True,
                        message=f"Message sent successfully to {request.chat_id}",
                        message_id=result.get('result', {}).get('message_id')
                    )
                else:
                    error_msg = result.get('description', 'Unknown error')
                    logger.error(f"Telegram API error: {error_msg}")
                    return TelegramSendResponse(
                        success=False,
                        message=f"Telegram API error: {error_msg}"
                    )

    except aiohttp.ClientError as e:
        logger.error(f"Telegram network error: {e}")
        return TelegramSendResponse(
            success=False,
            message=f"Network error: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Telegram send failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send message: {str(e)}")

@router.post("/telegram/test", response_model=TelegramSendResponse)
async def test_telegram_config(config: TelegramConfig):
    """
    Test Telegram bot configuration

    Verifies bot token by fetching bot information.
    """
    try:
        url = f"https://api.telegram.org/bot{config.bot_token}/getMe"

        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                result = await response.json()

                if result.get('ok'):
                    bot_info = result.get('result', {})
                    bot_name = bot_info.get('username', 'Unknown')
                    return TelegramSendResponse(
                        success=True,
                        message=f"Bot token is valid. Connected to @{bot_name}"
                    )
                else:
                    error_msg = result.get('description', 'Invalid token')
                    return TelegramSendResponse(
                        success=False,
                        message=f"Bot token invalid: {error_msg}"
                    )

    except Exception as e:
        logger.error(f"Telegram config test failed: {e}")
        return TelegramSendResponse(
            success=False,
            message=f"Connection failed: {str(e)}"
        )
